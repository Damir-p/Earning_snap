from docx import Document
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
import os

class Converter:
    def __init__(self, input_path):
        self.input_path = input_path
        self.output_path = self.generate_output_path()

    def generate_output_path(self):
        return self.input_path.rsplit('.', 1)[0] + '.pdf'

    def convert_to_pdf(self):
        if self.input_path.endswith('.docx'):
            self.convert_and_remove(self.convert_docx_to_pdf)
        elif self.input_path.endswith('.xlsx'):
            self.convert_and_remove(self.convert_xlsx_to_pdf)
        else:
            print("Неподдерживаемый формат файла")

    def convert_and_remove(self, conversion_function):
        conversion_function()
        os.remove(self.input_path)
        print(f"Исходный файл удален: {self.input_path}")

    def convert_docx_to_pdf(self):
        doc = Document(self.input_path)
        pdf_canvas = canvas.Canvas(self.output_path)
        for paragraph in doc.paragraphs:
            pdf_canvas.drawString(100, 800, paragraph.text)
            pdf_canvas.showPage()
        pdf_canvas.save()

    def convert_xlsx_to_pdf(self):
        wb = load_workbook(self.input_path)
        pdf_canvas = canvas.Canvas(self.output_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                pdf_canvas.drawString(100, 800, ' '.join(map(str, row)))
                pdf_canvas.showPage()
        pdf_canvas.save()

if __name__ == "__main__":
    input_path = input("Введите полный путь к файлу (с расширением .docx или .xlsx): ")
    converter = Converter(input_path)

    if not os.path.isfile(converter.input_path):
        print("Файл не найден. Убедитесь, что вы ввели правильный путь.")
    else:
        print(f"Выбранный файл: {converter.input_path}")
        converter.convert_to_pdf()
        print(f"Файл успешно сконвертирован в PDF. Результат сохранен по пути: {converter.output_path}")
